##########################################################################################################
#   Description: 针对业务进行专门的分析和处理
#   Authors:     BaiYuan <395642104@qq.com>
##########################################################################################################
from .router import PowerGridGraph
from .substation import Substation

from craftHub.tool import gLog

import os
import re
from pathlib import Path
from typing import Optional, List, Tuple, Literal

import openpyxl as ol
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont


class PartialChain:
    '''部分链路
    每一条部分链路将被标记为红色或者黑色'''
    def __init__(self, textBlock: TextBlock | str | Cell, preSta: Optional[str] = None, nextSta: Optional[str] = None):
        """部分链路， 每一个部分链路处理一个文本块或者一个str

        :param textBlock: 文本块信息, 可以直接输入cell, 直接输入cell时, 表明该cell不是富文本
        :param preSta:    前驱节点名字
        :param nextSta:   后驱节点名字
        """        
        
        self.pStaList: List[str] = []   # 节点列表
        self.pStartSta = None           # 部分链路开始节点
        self.pEndSta = None             # 部分链路结束节点
        self.preSta = preSta            # 前驱节点， 如果使用部分链路开始和结束节点无法计算出来，考虑使用前驱节点和后驱节点计算
        self.nextSta = nextSta          # 后驱节点
        self.color: Literal["000000", "FF0000"]  = "000000" # 颜色,红色或者黑色
        self.isEntered = False          # 是否被截断，用于截断链路初始化
        
        self.rawText: str = ""               # 通常文本信息
        
        # 文本块分类, 提取信息
        if isinstance(textBlock, str):
            self.color = "000000"
            self.rawText = textBlock
        elif isinstance(textBlock, TextBlock):
            self.color = "FF0000" if self._isRed(textBlock) else "000000"
            self.rawText = textBlock.text
        elif isinstance(textBlock, Cell):
            self.color = "FF0000" if self._isRed(textBlock) else "000000"
            self.rawText = str(textBlock.value)
        else:
            raise AttributeError(f"处理文本块遇到未知情况, 文本块类型为: {type(textBlock)}")
        
        # 截断字符串
        for c in ["\n", "(", "（", "光缆路由"]:
            if c in self.rawText:
                self.rawText = self.rawText.split(c)[0]
                self.isEntered = True
                break
            
        # 获取站点列表
        self.rawText = self.rawText.replace(" ", "")
        self.rawText = self.rawText.replace("——", "-")

        print(f"部分链路文本:{self.rawText}")
        if "->" in self.rawText:
            self.pStaList = [pSta for pSta in self.rawText.split("->") if pSta]  # 支持->格式
        elif "-" in self.rawText:
            self.pStaList = [pSta for pSta in self.rawText.split("-") if pSta]   # 支持-格式
        elif self.rawText:
            self.pStaList = [self.rawText]

        if len(self.pStaList) >= 1:
            self.pStartSta = self.pStaList[0]
            self.pEndSta = self.pStaList[-1]
            
        # 获取节点判断，可能无法获取成功
        if not self.pStartSta or not self.pEndSta:
            # gLog.logInfoWithNoTime(f"Warning: 获取部分链路开始节点或结束节点失败: {self.rawText}")
            pass
            
    def isEmpty(self):
        '''空列表判断'''
        return not bool(self.pStaList)
        
    def set(self, preSta: Optional[str] = None, nextSta: Optional[str] = None):
        """设置前驱节点和后驱节点信息

        :param preSta: 前驱节点
        :param nextSta: 后驱节点
        """        
        
        if preSta is not None:
            self.preSta = preSta
        if nextSta is not None:
            self.nextSta = nextSta
        
    @staticmethod
    def _isRed(textBlock: TextBlock | str | Cell):
        """红色文本判断, 严格判断字体是否是FF0000

        :param textBlock: 文本块
        """        
        try:
            # 检查是否有字体属性
            if hasattr(textBlock, 'font') and textBlock.font:
                font = textBlock.font
                # 检查颜色
                if hasattr(font, 'color') and font.color:

                    # 获取颜色值
                    colorValue = None
                    if hasattr(font.color, 'rgb'):
                        colorValue = font.color.rgb
                    elif hasattr(font.color, 'type'):
                        colorValue = font.color.type
                    
                    # 判断是否是红色
                    if colorValue and str(colorValue).upper().endswith("FF0000"):
                        return True
        except:
            pass
        return False

    def statics(self):
        '''显示自身信息'''
        if self.color == "FF0000":
            gLog.logInfoWithNoTime(f"{gLog.RED}\'{self.pStaList}\'{gLog.END}, 前驱节点: \'{self.preSta}\', 后驱节点: \'{self.nextSta}\'")
        else:
            gLog.logInfoWithNoTime(f"\'{self.pStaList}\', 前驱节点: \'{self.preSta}\', 后驱节点: \'{self.nextSta}\'")

    def calculate(self, graph: PowerGridGraph, 
                        forbiddenPaths: List[List[str]]
                        ):
        '''计算部分链路信息
        如果使用红色链路双边节点计算无结果, 将尝试使用前驱节点或者后驱节点计算

        :param graph:           电力网拓扑
        :param forbiddenPaths:  禁用路径
        :return: List[staName] 或者 List[Substation]
        '''      
        
        # 黑色链路，无需计算, 返回自身链路列表
        if self.color == "000000":
            return self.pStaList
        
        failReason = "未知错误" # 失败原因
        # 红色链路，先尝试使用红色双边节点计算
        try:
            _, minPath = graph.findShortestPath(startName = self.pStartSta,
                                                endName = self.pEndSta,
                                                forbiddenPaths = forbiddenPaths)
            return minPath
        except Exception as e:
            failReason = f"使用双边节点计算失败: {str(e)}"
            
        # 双边节点计算失败，尝试使用前驱节点计算
        if self.preSta:
            try:
                _, minPath = graph.findShortestPath(startName = self.preSta,
                                                    endName = self.pEndSta,
                                                    forbiddenPaths = forbiddenPaths)
                return minPath[1:]
            except Exception as e:
                failReason = f"使用前驱节点计算失败: {str(e)}"
                
        # 使用后驱节点计算
        if self.nextSta:
            try:
                _, minPath = graph.findShortestPath(startName = self.pStartSta,
                                                    endName = self.nextSta,
                                                    forbiddenPaths = forbiddenPaths)
                return minPath[:-1]
            except Exception as e:
                failReason = f"使用后驱节点计算失败: {str(e)}"
                
        # 同时使用前驱和后驱节点计算
        if self.preSta and self.nextSta:
            try:
                _, minPath = graph.findShortestPath(startName = self.preSta,
                                                    endName = self.nextSta,
                                                    forbiddenPaths = forbiddenPaths)
                return minPath[1:-1]
            except Exception as e:
                failReason = f"使用前驱节点和后驱节点计算失败: {str(e)}"

        raise ValueError(failReason)
    

class Chain:
    '''富文本或普通文本格式链路
    富文本链路具有颜色属性, 黑色属性链路应当保留， 红色属性链路应当重建'''
    def __init__(self, startSta: str, endSta: str, chainCell: Cell):
        """链路初始化

        :param startSta:    链路开始节点
        :param endSta:      链路结束节点
        :param chainCell:   链路单元格，可以从链路单元格获取文本信息，颜色信息和位置信息
        :raises Exception: _description_
        """

        self.startSta = startSta    # 链路开始节点
        self.endSta = endSta        # 链路结束节点
        self.chainCell = chainCell

        self.paritialChainList: List[PartialChain] = [] # 部分链路列表
        
        # 部分链路基本初始化
        if isinstance(chainCell.value, CellRichText):
            for textObj in chainCell.value:
                
                newPartialChain = PartialChain(textBlock = textObj,
                                               preSta = None,
                                               nextSta = None
                                               )
                self.paritialChainList.append(newPartialChain)
                
                # 回车键后续的信息不处理
                if newPartialChain.isEntered:
                    break;
        else:
            newPartialChain = PartialChain(textBlock = chainCell,
                                            preSta = None,
                                            nextSta = None
                                            )
            self.paritialChainList.append(newPartialChain)
                
        # 部分链路同色合并
        mergedPchainList: List[PartialChain] = []
        for curPchain in self.paritialChainList:
            # 合并
            try:
                mergedPchain = self.merge(mergedPchainList[-1], curPchain)
            except Exception as e:
                mergedPchainList.append(curPchain)
            else:
                mergedPchainList[-1] = mergedPchain

        self.paritialChainList = mergedPchainList

        # 空值过滤
        self.paritialChainList = [pc for pc in self.paritialChainList if not pc.isEmpty()]

        # 获取前驱节点和后驱节点
        for index in range(len(self.paritialChainList)):
            if index > 0:
                self.paritialChainList[index].set(preSta = self.paritialChainList[index - 1].pEndSta)
            if index < len(self.paritialChainList) - 1:
                self.paritialChainList[index].set(nextSta = self.paritialChainList[index + 1].pStartSta)
                

    @staticmethod
    def merge(pchain1: PartialChain, pchain2: PartialChain):
        """合并两个部分链路

        :param pchain1: 链路1
        :param pchain2: 链路2
        """        
        if pchain1.color != pchain2.color:
            raise ValueError("颜色不一致，无法合并")
        
        # 黑色链路合并
        if pchain1.color == "000000":
            newPchain = PartialChain(textBlock = pchain1.rawText + pchain2.rawText)
            return newPchain
        
        # 红色链路合并
        else:
            newPchain = PartialChain(textBlock = TextBlock(InlineFont(color = "FF0000"), text = pchain1.rawText + pchain2.rawText))
            return newPchain

    def isEmpty(self):
        '''空列表判断'''
        return not bool(self.paritialChainList)

    def statics(self):
        '''显示自身信息'''
        if self.isEmpty():
            gLog.logInfoWithNoTime(f"  {gLog.BLUE}空链路{gLog.END}")
        
        for index, pchain in enumerate(self.paritialChainList):
            gLog.logInfoWithNoTime(f"  部分链路\'{index}\': ", end = " ")
            pchain.statics()

    def getpStaList(self, color: Literal["FF0000", "000000"])->List[List[str]]:
        """获取部分站点列表, 一般用于禁用

        :param color:   可选红色或者黑色
        :return:        为列表[列表[str]]格式
        """        
        
        redpStaList: List[List[str]] = []     # 红色禁用列表， 整个业务有效
        blackpStaList: List[List[str]] = []   # 黑色禁用列表， 仅对一个链路有效，防止产生回路
        
        for pChain in self.paritialChainList:
            if pChain.color == "FF0000":
                redpStaList.append(pChain.pStaList)
            else:
                blackpStaList.append(pChain.pStaList)
                
        if color == "FF0000":
            return redpStaList
        else:
            return blackpStaList

    def calculate(self, graph: PowerGridGraph, 
                        forbiddenPaths: List[List[str]]):
        """计算完整链路信息, 
        将返回计算后得到的cell, 以及产生的新的红色路径

        :param graph:           电力网拓扑
        :param forbiddenPaths:  禁用路径
        :raises AssertionError: 开发出错时，抛出断言错误
        :return: CellRichText([TextBlock]), List[List[str]]
        """        

        textBlockList: List[TextBlock] = []                         # 文本块列表
        blackForbiddenPaths = self.getpStaList(color = "000000")    # 黑色禁用列表
        newRedForbiddenPaths: List[List[str]] = []                  # 新产生的红色路径
        
        # 部分链路非空，调用部分链路计算
        if self.paritialChainList:
            for index, pChain in enumerate(self.paritialChainList):           
                try:
                    result = pChain.calculate(graph = graph,
                                              forbiddenPaths = forbiddenPaths + blackForbiddenPaths + newRedForbiddenPaths)
                    if not result:
                        continue
                    
                    # 从计算结果获取文本块
                    if isinstance(result[0], str):
                        newTb = TextBlock(InlineFont(color = pChain.color), "->".join(result))
                    elif isinstance(result[0], Substation):
                        newRedForbiddenPaths.append([s.name for s in result])   # 新生成的红色路径
                        newTb = TextBlock(InlineFont(color = pChain.color), graph.pathToStr(result))
                    else:
                        raise AssertionError("部分链路计算结果不是字符串列表或变电站列表！")
            
                    textBlockList.append(newTb)
                    if index != len(self.paritialChainList) -1:
                        textBlockList.append(TextBlock(InlineFont(color = "000000"), "->"))  
                except AssertionError as e:
                    raise e
                except Exception as e:
                    raise Exception(f"第 {index} 个部分链路计算失败: {str(e)}")

        # 部分链路为空，直接调用链路开始节点和结束节点计算
        else:
            _, result = graph.findShortestPath(startName = self.startSta,
                                               endName = self.endSta,
                                               forbiddenPaths = forbiddenPaths)
            newTb = TextBlock(InlineFont(color = "FF0000"), graph.pathToStr(result))
            textBlockList.append(newTb)
                
        return CellRichText(textBlockList), newRedForbiddenPaths


class Business:
    '''一个业务
    一个业务一般在一个表格中占据两道三行的位置
    一个业务可能包含一个主用链路和两条备用链路'''
    
    patternStd = r'^(.*?)\n[（(]光缆路由[：:](.*?)[）)]$'
    def __init__(self, name: str, chainLists: Tuple[Cell]):
        self.name = name                    # 业务名
        self.chainNum = len(chainLists)     # 链路数量
        self.startSta: str = None           # 业务起点
        self.endSta: str = None             # 业务终点

        # 获取业务起点和终点(存在问题， 不完全可靠)
        for chainCell in chainLists:
            chainValue = str(chainCell.value)
            match = re.match(self.patternStd, chainValue)
            if match:
                opticalChain = match.group(1).strip()   # 光路路由字符串
                fiberChain = match.group(2).strip()     # 光缆路由字符串

                print(f"光路路由: {opticalChain}")

                if self.startSta is None and "->" in opticalChain:
                    if "->" in opticalChain:
                        self.startSta = opticalChain.split("->")[0].strip()
                        self.endSta = opticalChain.split("->")[-1].strip()
                        break
                    elif "-" in opticalChain:
                        self.startSta = opticalChain.split("-")[0].strip()
                        self.endSta = opticalChain.split("-")[-1].strip()
                        break
                
            elif "-" in chainValue and "光缆" not in chainValue:
                self.startSta = chainValue.split("-")[0].strip()
                self.endSta = chainValue.split("-")[-1].strip()
                break
                
        if self.startSta is None:
            raise ValueError(f"业务 \'{self.name}\' 无法获取业务起点和终点信息: ")
          
        # 初始化链路信息
        self.chains: List[Chain] = []
        for index, cell in enumerate(chainLists):
            try:
                newChain = Chain(self.startSta, self.endSta, cell)
                self.chains.append(newChain)
            except AssertionError as e:
                raise e
            except Exception as e:
                gLog.logInfo(f"业务 \'{self.name}\' 链路 \'{index}\' 初始化错误: {str(e)}")

    def statics(self):
        '''显示自身信息'''
        gLog.logInfoWithNoTime(f"业务 \'{self.name}\'")
        for index, chain in enumerate(self.chains) :
            gLog.logInfoWithNoTime(f"链路\'{index}\':")
            chain.statics()

    def calculate(self, graph: PowerGridGraph, forbiddenPaths: List[List[str]]):
        """计算本业务的信息
        返回cell列表和对应的位置列表

        :param graph:           电力网拓扑
        :param forbiddenPaths:  禁用路径, 应当是全局禁用路径

        :raises AssertionError: 开发出错时，抛出断言错误
        :return: Tuple(List[CellRichText], List[str])
        返回构建完成的富文本列表和他们对应的坐标列表
        """    
        cellRichTextList: List[CellRichText] = []       # 创建的cell列表，是富文本格式
        coordinateList: List[str] = []                  # 坐标列表，用于写出数据
        redForbiddenPaths: List[List[str]] = []         # 红色禁用路径，整个业务内有效

        for chain in self.chains:
            redForbiddenPaths += chain.getpStaList(color = "FF0000")

        for chain in self.chains:
            try:
                newCell, newRedForbiddenPaths = chain.calculate(graph = graph,
                                                                forbiddenPaths = forbiddenPaths + redForbiddenPaths)
                redForbiddenPaths += newRedForbiddenPaths   # 更新红色禁用
            except AssertionError as e:
                raise e
            except Exception as e:
                newCell = CellRichText([TextBlock(InlineFont(color = "0000FF"), str(e))])
            finally:
                coordinate = f"{chr(ord(chain.chainCell.column_letter) + 1)}{chain.chainCell.row}" # 变换获得右边一格坐标
                cellRichTextList.append(newCell)
                coordinateList.append(coordinate)
                
        return cellRichTextList, coordinateList
    

class Businesses:
    '''业务群
    业务群从一个xlsx文件初始化, 仅计算其中一个sheet
    一个业务群将包含复数各业务'''
    def __init__(self, filePath: Path):
        """业务群初始化

        :param filePath:    文件路径，必须是.xlsx
        """    
        
        # 检查    
        if not filePath.exists():
            raise FileNotFoundError(f"业务文件未找到: {filePath}")
        if filePath.suffix != ".xlsx":
            raise ValueError(f"业务文件只支持 \'.xlsx\' 格式， \'{filePath.suffix}\' 不支持")
        
        self.wb: Workbook = ol.load_workbook(filename = str(filePath),
                                             read_only = False,
                                             rich_text = True)
        
        self.ws = self.wb.active                # 获取当前活动表单
        self.businessList: List[Business] = []  # 业务列表
        self.colNameDict = {}                   # 名字映射列名字典

        for cell in self.ws[1]:
            self.colNameDict[str(cell.value)] = cell.column_letter
            
        # 列名检查
        needColumns = ["电路名称", "承载网络", "路由类别", "当前路由", "调整后路由"]
        for col in needColumns:
            if col not in self.colNameDict:
                raise KeyError(f"格式匹配失败，业务表中没有找到列 \'{col}\'")

        colCircuitName = self.ws[self.colNameDict["电路名称"]][1:]
        colCurRoute = self.ws[self.colNameDict["当前路由"]][1:]
        
        # 创建业务列表
        currentName = ""
        currentChainLists: List[Cell] = []
        failedCount = 0
        for index, (name, curChain) in enumerate(zip(colCircuitName, colCurRoute)):
            # 记录名字, 添加列
            currentChainLists.append(curChain)
            if name.value is not None:
                currentName = str(name.value)
            
            # 如果到达末尾或者下个名字非空，创建业务
            if index == len(colCircuitName) - 1 or colCircuitName[index + 1].value is not None:
                try:
                    newBusiness = Business(name = currentName,
                                        chainLists = currentChainLists)
                except Exception as e:
                    gLog.logInfo(f"Warning: 对于业务的初始化失败: {str(e)}")
                    failedCount += 1
                    continue
                else:
                    self.businessList.append(newBusiness)
                finally:
                    currentName = ""
                    currentChainLists.clear()

        gLog.logInfo(f"{len(self.businessList)} 个业务初始化成功，{failedCount} 个业务初始化失败")

    def statics(self):
        '''显示自身信息'''
        for b in self.businessList:
            b.statics()
            gLog.enter()
        gLog.logInfoWithNoTime(f"共计 \'{len(self.businessList)}\' 个业务")

    def calculate(self, graph: PowerGridGraph, 
                        forbiddenPaths: List[List[str]],
                        savePath: Path):
        """计算全部业务的信息，并尝试填表

        :param graph:           电力网拓扑
        :param forbiddenPaths:  禁用路径
        :param savePath:  存储路径
        :raises AssertionError: 开发出错时，抛出断言错误
        :return: Tuple(List[CellRichText], List[str])
        返回构建完成的富文本列表和他们对应的坐标列表
        """    

        cellRichTextListAll: List[CellRichText] = []    # 全部的富文本cell构成的列表
        coordinateListAll: List[str] = []               # 全部的坐标构成的列表

        # 提取全部业务信息
        for index, business in enumerate(self.businessList) :
            gLog.logInfo(f"{index}.处理业务 \'{business.name}\'")
            cellRichTextList, coordinateList = business.calculate(graph = graph,
                                                                forbiddenPaths = forbiddenPaths)
            cellRichTextListAll += cellRichTextList
            coordinateListAll += coordinateList
            
        gLog.logInfo(f"获取到计算结果 {len(cellRichTextListAll)} 个")
            
        # 写出全部业务信息
        if False:    
            for cellRichText, coordinate in zip(cellRichTextListAll, coordinateListAll):
                self.ws[coordinate].value = cellRichText
                    
            # 保存文件
            self.wb.save(str(savePath))
            gLog.logInfo(f"计算结果已保存到: {savePath}")
            os.startfile(str(savePath))
        else:
            newWb = Workbook()
            newWs = newWb.active
            for cellRichText, coordinate in zip(cellRichTextListAll, coordinateListAll):
                newWs[coordinate].value = cellRichText
            newWb.save(str(savePath))
            gLog.logInfo(f"计算结果已保存到: {savePath}")
            os.startfile(str(savePath))

